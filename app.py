import os, json, time, tempfile, requests
from flask import Flask, request, jsonify, send_from_directory, redirect, session
from urllib.parse import urlencode
import openpyxl, anthropic

app = Flask(__name__, static_folder="static")
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "okto-sync-secret-2024-xK9p")

ANTHROPIC_API_KEY   = os.environ.get("ANTHROPIC_API_KEY", "")
AZURE_CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID", "c4875a3b-811a-428b-8a34-a492257db518")
AZURE_CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")
AZURE_TENANT_ID     = os.environ.get("AZURE_TENANT_ID", "7e4f0df0-f07f-4934-97a9-872d0c5292f9")
REDIRECT_URI        = os.environ.get("REDIRECT_URI", "https://okto-sync-production.up.railway.app/auth/callback")
SCOPES              = "openid profile email Files.Read.All Sites.Read.All User.Read offline_access"

ASANA_FIELDS = {
    "market_value":"1209691181990947","debt":"1209690936207910","noi":"1209690933762869",
    "dscr":"1209691184219184","ltv":"1209691185448001","cap_rate":"1209691329150592",
    "cost":"1209693994446597","units":"1209720926990044","sf_building":"1209720821050620",
    "sf_land":"1209759849844611","gross_income":"1209762019720933","cashflow":"1210006419611186",
    "expense":"1210006388018561","cashdown":"1210333099357386","price":"1209636012311297",
    "cash_call":"1210006382272158",
}

# Full project map with MANY aliases so fuzzy matching always works
PROJECT_GIDS = {
    # Calgary
    "Multi-Res IPP - Calgary - Riverview":"1210853284875107",
    "Calgary Riverview":"1210853284875107","Riverview":"1210853284875107","Calgary":"1210853284875107",
    # Mangin
    "Land - NCC Mangin Land":"1209331149696129","Mangin":"1209331149696129","NCC Mangin":"1209331149696129",
    # Beaumont
    "Ind IPP - 700 Avenue Beaumont":"1209331149696124","Beaumont":"1209331149696124","700 Beaumont":"1209331149696124","700 Avenue Beaumont":"1209331149696124",
    # Edmonton Garneau
    "Multi-Res IPP - Edmonton - Garneau":"1209331149696094","Garneau":"1209331149696094","Edmonton Garneau":"1209331149696094",
    # Edmonton Imperial
    "Multi-Res IPP - Edmonton - Imperial":"1209556797269329","Imperial":"1209556797269329","Edmonton Imperial":"1209556797269329",
    # Edmonton Insignia
    "Multi-Res IPP - Edmonton - Insignia":"1209624097039783","Insignia":"1209624097039783","Edmonton Insignia":"1209624097039783","Okto Insignia":"1209624097039783",
    # Edmonton Capilano
    "Multi-Res IPP - Edmonton - Capilano Tower":"1209331149696099","Capilano":"1209331149696099","Capilano Tower":"1209331149696099","Edmonton Capilano":"1209331149696099",
    # Allumetieres
    "Multi-Res - Allumetieres":"1209331149696158","Allumetieres":"1209331149696158","Allumettières":"1209331149696158",
    # Place Laval
    "Multi-Res - Place Laval Land":"1209331149696104","Place Laval":"1209331149696104","Laval Land":"1209331149696104","Laval":"1209331149696104",
    # Edmonton Lansdowne
    "Multi-Res IPP - Edmonton - Lansdowne":"1209771721895696","Lansdowne":"1209771721895696","Edmonton Lansdowne":"1209771721895696",
    # Edmonton Galbraith
    "Multi-Res IPP - Edmonton - Galbraith":"1209771721895687","Galbraith":"1209771721895687","Edmonton Galbraith":"1209771721895687",
    # Edmonton Axcess
    "Multi-Res IPP - Edmonton - Axcess":"1209198438088491","Axcess":"1209198438088491","Edmonton Axcess":"1209198438088491",
    # Edmonton Hamptons
    "Multi-Res IPP - Edmonton - Hamptons":"1209198438088484","Hamptons":"1209198438088484","Edmonton Hamptons":"1209198438088484",
    # Gatineau Eleonore
    "Multi-Res IPP - Gatineau - Eleonore":"1203839657006504","Eleonore":"1203839657006504","Éléonore":"1203839657006504","Gatineau Eleonore":"1203839657006504",
    # Alexandra
    "Multi-Res - Alexandra":"1203878731582881","Alexandra":"1203878731582881",
    # Dorval Walt
    "Multi-Res IPP - Dorval - Walt":"1203878731582875","Walt":"1203878731582875","Dorval Walt":"1203878731582875","Dorval":"1203878731582875",
    # Moqueurs
    "Multi-Res - Moqueurs":"1203878731582869","Moqueurs":"1203878731582869","Rue des Moqueurs":"1203878731582869",
    # 720 MTL-TO
    "Multi-Res - 720 MTL-TO (Romeo)":"1209120065416144","720 MTL-TO":"1209120065416144",
    "720 MTL TO":"1209120065416144","720 Montreal-Toronto":"1209120065416144",
    "720 MTL-TO (Romeo)":"1209120065416144","Romeo":"1209120065416144","Roméo":"1209120065416144",
    "720":"1209120065416144","MTL-TO":"1209120065416144",
    # 605 Wilfrid-Hamel
    "Multi-Res - 605 Wilfrid-Hamel (Charles)":"1203878731582893","605 Wilfrid-Hamel":"1203878731582893",
    "Wilfrid-Hamel":"1203878731582893","Charles":"1203878731582893","605":"1203878731582893",
    # 4717 Wellington
    "Multi-Res - 4717 Wellington":"1209671658434874","Wellington":"1209671658434874","4717 Wellington":"1209671658434874",
    # 311 Eglise
    "Multi-Res - 311 de l'Eglise":"1209671805092934","311 Eglise":"1209671805092934","de l'Église":"1209671805092934","311":"1209671805092934",
    # Loblaws Cowansville
    "Commercial - Loblaws Cowansville":"1210023255872586","Cowansville":"1210023255872586","Loblaws Cowansville":"1210023255872586","Loblaws Cowan":"1210023255872586",
    # Sherbrooke BB
    "Storage - 12455 Sherbrooke Est (BB)":"1206322186668459","Sherbrooke":"1206322186668459","12455 Sherbrooke":"1206322186668459","Sherbrooke BB":"1206322186668459",
    # Laurier
    "Land - Laurier":"1209120065416220","Laurier":"1209120065416220","NCC Laurier":"1209120065416220",
    # Acton Vale
    "Commercial - Loblaws Acton Vale":"1203831185421665","Acton Vale":"1203831185421665","Loblaws Acton Vale":"1203831185421665","Acton":"1203831185421665",
    # 1400 Taschereau
    "Commercial - 1400 Taschereau":"1209569869960748","Taschereau":"1209569869960748","1400 Taschereau":"1209569869960748",
    # CineParc
    "Land - Land CineParc":"1203878731582887","CineParc":"1203878731582887","Ciné-Parc":"1203878731582887","Cine Parc":"1203878731582887",
    # Walkley
    "Land - Land Walkley":"1203831300537692","Walkley":"1203831300537692","Mokto Walkley":"1203831300537692","Land Walkley":"1203831300537692",
    # Couture
    "Ind IPP - Couture":"1209198438088461","Couture":"1209198438088461","Bd Couture":"1209198438088461",
    # 7725 Cordner
    "Ind IPP - 7725 Cordner":"1206343135145159","Cordner":"1206343135145159","7725 Cordner":"1206343135145159",
    # Chateauguay
    "Ind IPP - Chateauguay 315 BLVD Industrial":"1209331149696153","Chateauguay":"1209331149696153","315 BLVD Industrial":"1209331149696153","315 Blvd":"1209331149696153",
    # Hamel
    "Ind IPP - Hamel":"1207216084898282","Hamel":"1207216084898282","Wilfrid Hamel":"1207216084898282","Ind Hamel":"1207216084898282",
    # Newton
    "Ind IPP - Newton":"1209785685167226","Newton":"1209785685167226","Av Newton":"1209785685167226",
    # Charest BB
    "Storage - Charest (BB)":"1203877426888063","Charest":"1203877426888063","Charest BB":"1203877426888063","Boulevard Charest":"1203877426888063",
    # Seigneuriale BB
    "Storage - Seigneuriale (BB)":"1203877426888066","Seigneuriale":"1203877426888066","Seigneuriale BB":"1203877426888066",
    # Leman BB
    "Storage - Leman, Laval (BB)":"1203877426888084","Leman":"1203877426888084","Leman Laval":"1203877426888084","Leman BB":"1203877426888084",
}

def fuzzy_match_project(name):
    """Try to match a project name to a GID using multiple strategies."""
    if not name: return None
    # 1. Exact match
    if name in PROJECT_GIDS: return PROJECT_GIDS[name]
    # 2. Case-insensitive exact
    nl = name.lower()
    for k,v in PROJECT_GIDS.items():
        if k.lower()==nl: return v
    # 3. Substring match — name contains key or key contains name
    for k,v in PROJECT_GIDS.items():
        if nl in k.lower() or k.lower() in nl: return v
    # 4. Word overlap — any significant word matches
    words = set(nl.split()) - {'the','a','an','and','or','of','in','de','la','le','les','du'}
    for k,v in PROJECT_GIDS.items():
        kwords = set(k.lower().split()) - {'the','a','an','and','or','of','in','de','la','le','les','du'}
        if words & kwords: return v
    return None

# ── Auth ──────────────────────────────────────────────────────────────
@app.route("/auth/login")
def auth_login():
    params={"client_id":AZURE_CLIENT_ID,"response_type":"code","redirect_uri":REDIRECT_URI,"scope":SCOPES,"response_mode":"query"}
    return redirect(f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/authorize?{urlencode(params)}")

@app.route("/auth/callback")
def auth_callback():
    code=request.args.get("code")
    if not code: return redirect("/?error=no_code")
    resp=requests.post(f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token",data={
        "client_id":AZURE_CLIENT_ID,"client_secret":AZURE_CLIENT_SECRET,
        "code":code,"redirect_uri":REDIRECT_URI,"grant_type":"authorization_code","scope":SCOPES,
    })
    tokens=resp.json()
    if "error" in tokens: return redirect("/?error=auth_failed")
    session["access_token"]=tokens.get("access_token")
    me=requests.get("https://graph.microsoft.com/v1.0/me",headers={"Authorization":f"Bearer {session['access_token']}"}).json()
    session["user_name"]=me.get("displayName","User")
    session["user_email"]=me.get("mail") or me.get("userPrincipalName","")
    return redirect("/")

@app.route("/auth/logout")
def auth_logout():
    session.clear(); return redirect("/")

@app.route("/auth/me")
def auth_me():
    if "access_token" not in session: return jsonify({"logged_in":False})
    return jsonify({"logged_in":True,"name":session.get("user_name"),"email":session.get("user_email")})

# ── SharePoint ────────────────────────────────────────────────────────
def sp_search(query, access_token):
    r=requests.post("https://graph.microsoft.com/v1.0/search/query",
        headers={"Authorization":f"Bearer {access_token}","Content-Type":"application/json"},
        json={"requests":[{"entityTypes":["driveItem"],"query":{"queryString":f"{query} filetype:xlsx"},
            "fields":["name","webUrl","lastModifiedDateTime","parentReference"],
            "from":0,"size":5,"sortProperties":[{"name":"lastModifiedDateTime","isDescending":True}]}]})
    results=[]
    try:
        for hit in r.json()["value"][0]["hitsContainers"][0]["hits"]:
            res=hit.get("resource",{}); pr=res.get("parentReference",{})
            results.append({"name":res.get("name",""),"webUrl":res.get("webUrl",""),
                "driveId":pr.get("driveId",""),"itemId":res.get("id",""),"modified":res.get("lastModifiedDateTime","")})
    except: pass
    return results

def read_excel_from_sp(drive_id, item_id, access_token):
    r=requests.get(f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization":f"Bearer {access_token}"},allow_redirects=True)
    if r.status_code!=200: return None
    with tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False) as tmp:
        tmp.write(r.content); tmp_path=tmp.name
    try:
        wb=openpyxl.load_workbook(tmp_path,data_only=True); out=[]
        for sn in wb.sheetnames:
            ws=wb[sn]; out.append(f"=== {sn} ===")
            for row in ws.iter_rows(values_only=True):
                if any(v is not None for v in row):
                    out.append("\t".join([str(v) if v is not None else "" for v in row]))
        return "\n".join(out)
    except: return None
    finally: os.unlink(tmp_path)

# ── Claude ────────────────────────────────────────────────────────────
def claude_call(system_prompt, user_content):
    client=anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    resp=client.messages.create(model="claude-sonnet-4-5",max_tokens=1200,
        system=system_prompt,messages=[{"role":"user","content":user_content}])
    raw=resp.content[0].text.strip()
    if raw.startswith("```"): raw=raw.split("```")[1]; raw=raw[4:] if raw.startswith("json") else raw
    return json.loads(raw.strip())

def build_system_prompt(extra=""):
    all_names = sorted(set(PROJECT_GIDS.keys()))
    # Give Claude the canonical names AND all aliases
    project_list = "\n".join(f"- {n}" for n in all_names)
    field_list   = "\n".join(f"- {n}" for n in ASANA_FIELDS)
    return f"""You are an Asana update assistant for Okto, a real estate investment firm based in Montreal/Gatineau.

You update project custom fields in Asana based on user instructions.

AVAILABLE PROJECTS (use the exact name OR any recognizable alias — the system will fuzzy-match):
{project_list}

AVAILABLE FIELDS:
{field_list}

FIELD ALIASES (understand these naturally):
- market value, valeur marchande, appraised value → market_value
- dette, loan balance, mortgage → debt  
- net operating income, revenu net → noi
- cap rate, taux de capitalisation, capitalization rate → cap_rate
- loan to value, ratio prêt-valeur → ltv
- debt service coverage, ratio de couverture → dscr
- coût, acquisition cost, total cost → cost
- square feet building, superficie bâtiment, pi² → sf_building
- square feet land, superficie terrain → sf_land
- revenu brut, gross revenue → gross_income
- flux de trésorerie, cash flow → cashflow
- dépenses, operating expenses → expense
- mise de fonds, down payment, equity → cashdown
- prix d'achat, purchase price → price
- appel de capital, capital call amount → cash_call
- nb unités, number of units, logements → units

Return ONLY valid JSON:
{{"updates":[{{"name":"project name (can be approximate)","fields":{{"field_name":numeric_value}},"summary":"human description"}}],"message":"fallback message","needs_sharepoint":true_or_false,"sharepoint_query":"search query if needed"}}

IMPORTANT:
- Accept French or English instructions
- Match project names loosely — "Hamel", "720", "Hamptons", "Walkley" etc. all work
- If user says "from SharePoint" or doesn't give a number, set needs_sharepoint=true
- Always return at least one update if you can identify any project and field
- Numbers can be written as "95M", "95 million", "95,000,000" — convert to numeric
{extra}"""

# ── Push to Asana ─────────────────────────────────────────────────────
def push_asana(updates, asana_token):
    results=[]
    for u in updates:
        name=u.get("name"); fields=u.get("fields",{}); summary=u.get("summary","")
        # Use fuzzy matching
        gid=fuzzy_match_project(name)
        if not gid:
            results.append({"name":name,"status":"skipped","reason":f"Could not match '{name}' to any project","summary":summary})
            continue
        # Get canonical name for display
        canonical=next((k for k,v in PROJECT_GIDS.items() if v==gid and " - " in k), name)
        cf={ASANA_FIELDS[fn]:float(v) for fn,v in fields.items() if fn in ASANA_FIELDS}
        if not cf:
            results.append({"name":canonical,"status":"skipped","reason":"No valid fields found","summary":summary})
            continue
        r=requests.put(f"https://app.asana.com/api/1.0/projects/{gid}",
            headers={"Authorization":f"Bearer {asana_token}","Content-Type":"application/json"},
            json={"data":{"custom_fields":cf}},timeout=15)
        time.sleep(0.2)
        if r.status_code==200:
            results.append({"name":canonical,"status":"success","summary":summary})
        else:
            try: err=r.json().get("errors",[{}])[0].get("message",r.text)
            except: err=r.text
            results.append({"name":canonical,"status":"error","reason":err,"summary":summary})
    return results

# ── Routes ────────────────────────────────────────────────────────────
@app.route("/")
def index(): return send_from_directory("static","index.html")

@app.route("/chat",methods=["POST"])
def chat():
    data=request.json; message=data.get("message","").strip(); asana_token=data.get("asana_token","").strip()
    if not message: return jsonify({"error":"No message"}),400
    if not asana_token: return jsonify({"error":"Asana token required"}),400
    if not ANTHROPIC_API_KEY: return jsonify({"error":"Anthropic key not configured"}),500

    try:
        parsed=claude_call(build_system_prompt(), message)
    except Exception as e:
        return jsonify({"error":f"AI error: {e}"}),500

    source_files=[]
    if parsed.get("needs_sharepoint"):
        if "access_token" not in session:
            return jsonify({"error":"sign_in_required","message":"Sign in with Microsoft to search SharePoint automatically"})
        query=parsed.get("sharepoint_query", message)
        files=sp_search(query, session["access_token"])
        if files:
            contents,names=[],[]
            for f in files[:3]:
                c=read_excel_from_sp(f["driveId"],f["itemId"],session["access_token"])
                if c: contents.append(c); names.append(f["name"])
            if contents:
                files_text="".join(f"\n\n=== FILE: {n} ===\n{c[:6000]}" for n,c in zip(names,contents))
                extra_prompt=f"\n\nSharePoint files found:\n{files_text}\n\nExtract the relevant numbers from these files."
                try:
                    parsed=claude_call(build_system_prompt(extra_prompt), f"User asked: {message}\n\nExtract numbers from the SharePoint files above and return updates.")
                    source_files=names
                except Exception as e:
                    return jsonify({"error":f"SharePoint read error: {e}"}),500

    updates=parsed.get("updates",[])
    if not updates:
        return jsonify({"updates":[],"message":parsed.get("message","I couldn't understand that. Try: 'update [project name] [field] to [value]'")})

    results=push_asana(updates, asana_token)
    return jsonify({"updates":results,"source_files":source_files})

@app.route("/analyze",methods=["POST"])
def analyze():
    if "file" not in request.files: return jsonify({"error":"No file"}),400
    file=request.files["file"]
    if not ANTHROPIC_API_KEY: return jsonify({"error":"Anthropic key not configured"}),500
    with tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False) as tmp:
        file.save(tmp.name)
        try:
            wb=openpyxl.load_workbook(tmp.name,data_only=True); out=[]
            for sn in wb.sheetnames:
                ws=wb[sn]; out.append(f"=== {sn} ===")
                for row in ws.iter_rows(values_only=True):
                    if any(v is not None for v in row):
                        out.append("\t".join([str(v) if v is not None else "" for v in row]))
            excel_text="\n".join(out)
            system=build_system_prompt()+'\n\nReturn JSON: {"projects":[{"name":"project","fields":{"field":value},"confidence":"high/medium/low","notes":"desc"}]}'
            parsed=claude_call(system, excel_text[:15000])
            # Normalize key
            if "updates" in parsed and "projects" not in parsed:
                parsed["projects"]=[{"name":u["name"],"fields":u.get("fields",{}),"confidence":"high","notes":u.get("summary","")} for u in parsed["updates"]]
            return jsonify(parsed)
        except Exception as e: return jsonify({"error":str(e)}),500
        finally: os.unlink(tmp.name)

@app.route("/sync",methods=["POST"])
def sync():
    data=request.json; asana_token=data.get("asana_token","").strip(); projects=data.get("projects",[])
    if not asana_token: return jsonify({"error":"Asana token required"}),400
    updates=[{"name":p.get("name"),"fields":p.get("fields",{}),"summary":""} for p in projects if p.get("name") and p.get("fields")]
    results=push_asana(updates, asana_token)
    return jsonify({"results":results})

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port)
